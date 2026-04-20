import io
import cv2
import requests
import openpyxl
import psycopg2.extras
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, render_template, jsonify, request, Response, send_file, make_response
from requests_toolbelt import MultipartEncoder
from datetime import datetime, timedelta, date

from config import get_faciais_conn, release_faciais_conn, get_pg_conn, release_pg_conn, HEIMDALL_URL, HEIMDALL_IMAGE_BASE, HEIMDALL_START_DATE, HEIMDALL_END_DATE, ZIONS_API_URL, ZIONS_TOKEN
from tracer import trace


def _carregar_cameras():
    """Carrega câmeras válidas com aliases compatíveis com o código legado."""
    try:
        conn = get_faciais_conn()
        try:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute("""
                SELECT
                    c.camera_id                     AS id_camera,
                    c.camera_name                   AS camera,
                    c.rtsp_url                      AS rstp,
                    c.created_at                    AS camera_criada_em,
                    ct.camera_type_id               AS id_tipo_camera,
                    ct.camera_type_name             AS tipo_camera,
                    s.store_id                      AS id_local,
                    s.cep,
                    s.address_number                AS numero,
                    s.address_complement            AS complemento,
                    co.company_id                   AS id_empresa,
                    co.company_name                 AS empresa,
                    cg.company_group_id             AS id_grupo,
                    cg.company_group_name           AS grupo,
                    cot.company_type_id             AS id_tipo_empresa,
                    cot.company_type_name           AS tipo_empresa,
                    rg.retailer_group_id            AS id_dono,
                    rg.retailer_group_name          AS dono
                FROM cameras c
                LEFT JOIN camera_types  ct  ON ct.camera_type_id  = c.camera_type_id
                LEFT JOIN stores        s   ON s.store_id          = c.store_id
                LEFT JOIN companies     co  ON co.company_id       = s.company_id
                LEFT JOIN company_groups cg ON cg.company_group_id = co.company_group_id
                LEFT JOIN company_types  cot ON cot.company_type_id = co.company_type_id
                LEFT JOIN retailer_groups rg ON rg.retailer_group_id = s.retailer_group_id
                ORDER BY c.camera_id
            """)
            rows = cursor.fetchall()
            cursor.close()
            conn.rollback()
        finally:
            release_faciais_conn(conn)
        ids = [str(r["id_camera"]) for r in rows if r.get("id_camera") is not None]
        return ids, rows
    except Exception as e:
        print(f"[tracks] Erro ao carregar câmeras do banco: {e}")
        return [], []


# IDs das câmeras a consultar e lista completa carregados da view vw_cameras_completo
CAMERA_IDS, CAMERAS_COMPLETO = _carregar_cameras()

tracks_bp = Blueprint("tracks", __name__)


def get_last_track_ids(limit=5):
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """SELECT track_id,
                      detection_record_id AS id,
                      created_at,
                      person_id           AS id_unico,
                      image_path
               FROM detection_records
               ORDER BY detection_record_id DESC LIMIT %s""",
            (limit,),
        )
        rows = cursor.fetchall()
        cursor.close()
        conn.rollback()
        return rows
    finally:
        release_faciais_conn(conn)


def query_heimdall(track_id):
    trace(track_id, f"query_heimdall: POST {HEIMDALL_URL} (câmeras: {CAMERA_IDS})")
    try:
        fields = [
            ("track_id",   track_id),
            ("confidence", "0.7"),
            ("start_date", HEIMDALL_START_DATE),
            ("end_date",   HEIMDALL_END_DATE),
        ]
        for cam_id in CAMERA_IDS:
            fields.append(("camera_id", cam_id))

        m = MultipartEncoder(fields=fields)
        resp = requests.post(
            HEIMDALL_URL,
            data=m,
            headers={"Content-Type": m.content_type},
            timeout=15,
        )
        if resp.status_code == 404:
            trace(track_id, "query_heimdall: track_id não encontrado no Heimdall (404) — sem matches")
            return {}, None
        if not resp.ok:
            trace(track_id, f"query_heimdall: ERRO HTTP {resp.status_code}")
            return None, f"HTTP {resp.status_code} — {resp.text[:500]}"
        trace(track_id, f"query_heimdall: resposta OK ({resp.status_code})")
        return resp.json(), None
    except requests.exceptions.RequestException as e:
        trace(track_id, f"query_heimdall: ERRO de rede → {e}")
        return None, str(e)
    except ValueError as e:
        trace(track_id, f"query_heimdall: JSON inválido → {e}")
        return None, f"JSON inválido: {e}"


def build_image_url(camera_id, filename):
    if not camera_id or not filename:
        return None
    return f"{HEIMDALL_IMAGE_BASE}/{camera_id}/{filename}"


@tracks_bp.route("/tracks")
def tracks_page():
    rows = get_last_track_ids(5)
    results = []
    for row in rows:
        track_id = row["track_id"]
        trace(track_id,"Chamando query_heimdall a partir de tracks_page")
        data, error = query_heimdall(track_id)
        results.append({
            "track_id": track_id,
            "db_id": row.get("id"),
            "db_created_at": str(row.get("created_at", "")),
            "data": data,
            "error": error,
            "image_base": HEIMDALL_IMAGE_BASE,
        })
    return render_template("tracks.html", results=results)


def fmt_timestamp(value):
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(str(value))
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    except (ValueError, TypeError):
        return value


def fmt_score(value):
    try:
        return f"{float(value) * 100:.1f}%"
    except (ValueError, TypeError):
        return value


GENDER_MAP = {0: "F", 1: "M"}


def get_best_face(track_id, data=None):
    """Retorna (image_path, camera_id, face_det_score) da face com maior score do Heimdall."""
    if data is None:
        trace(track_id, "get_best_face: iniciado e já chamando query_heimdall")
        data, error = query_heimdall(str(track_id))
        if error or not data:
            trace(track_id, f"get_best_face: sem dados do Heimdall → {error}")
            return None, None, None
    else:
        trace(track_id, "get_best_face: iniciado com dados já disponíveis")
    caras = (data.get("track") or {}).get("faces", [])
    trace(track_id, f"get_best_face: {len(caras)} face(s) encontrada(s) em track.faces")
    best = None
    best_score = -1
    for face in caras:
        if not face.get("image_path"):
            continue
        try:
            score = float(face.get("face_det_score") or 0)
        except (ValueError, TypeError):
            score = 0
        if score > best_score:
            best_score = score
            best = face
    if best:
        # face_recgn_score não existe em track.faces[] — está em matches[].
        # Usa o maior score entre todos os matches disponíveis.
        recgn_score = None
        for match in (data.get("matches") or []):
            try:
                s = float(match.get("face_recgn_score") or 0) or None
            except (ValueError, TypeError):
                s = None
            if s is not None and (recgn_score is None or s > recgn_score):
                recgn_score = s
        trace(track_id, f"get_best_face: melhor face → image_path={best['image_path']} camera_id={best.get('camera_id')} score={best_score:.4f} recgn={recgn_score}")
        return best["image_path"], best.get("camera_id"), best_score if best_score >= 0 else None, recgn_score
    trace(track_id, "get_best_face: nenhuma face com image_path encontrada")
    return None, None, None, None


def get_pessoas_by_ids(id_unicos):
    """Retorna dict {id_unico: {nome, apelido}} para os ids fornecidos."""
    ids = [i for i in id_unicos if i is not None]
    if not ids:
        return {}
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        placeholders = ",".join(["%s"] * len(ids))
        cursor.execute(
            f"""SELECT person_id AS id_unico, full_name AS nome, nickname AS apelido
                FROM people WHERE person_id IN ({placeholders})""",
            ids,
        )
        pessoas = {p["id_unico"]: p for p in cursor.fetchall()}
        cursor.close()
        conn.rollback()
        return pessoas
    finally:
        release_faciais_conn(conn)


@tracks_bp.route("/tracks/resumo")
def tracks_resumo():
    rows = get_last_track_ids(30)

    id_unicos = [row.get("id_unico") for row in rows]
    pessoas = get_pessoas_by_ids(id_unicos)

    results = []
    for row in rows:
        track_id = row["track_id"]
        data, error = query_heimdall(track_id)

        faces = []
        if data and isinstance(data, dict):
            matches = data.get("matches", [])
            for face in matches:
                score_raw = face.get("face_det_score")
                try:
                    if score_raw is None or float(score_raw) < 0.73:
                        continue
                except (ValueError, TypeError):
                    continue
                gender_raw = face.get("gender")
                recgn_raw = face.get("face_recgn_score")
                faces.append({
                    "id":               face.get("id"),
                    "track_id":         face.get("track_id", track_id),
                    "age":              face.get("age"),
                    "gender":           GENDER_MAP.get(gender_raw, gender_raw),
                    "face_det_score":   fmt_score(score_raw) if score_raw is not None else None,
                    "face_recgn_score": fmt_score(recgn_raw) if recgn_raw is not None else None,
                    "image_path":       face.get("image_path"),
                    "image_url":        HEIMDALL_IMAGE_BASE + face["image_path"] if face.get("image_path") else None,
                    "timestamp":        fmt_timestamp(face.get("timestamp")),
                })

        id_unico = row.get("id_unico")
        pessoa = pessoas.get(id_unico, {})
        results.append({
            "track_id":   track_id,
            "db_id":      row.get("id"),
            "id_unico":   id_unico,
            "nome":       pessoa.get("nome"),
            "apelido":    pessoa.get("apelido"),
            "image_path": row.get("image_path"),
            "error":      error,
            "faces":      faces,
        })
    return render_template("tracks_resumo.html", results=results)


@tracks_bp.route("/tracks/lista")
def tracks_lista():
    per_page = 5
    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1

    ids_raw = request.args.get('ids', '').strip()
    ids_filter = []
    if ids_raw:
        for part in ids_raw.replace('\n', ',').replace(';', ',').split(','):
            part = part.strip()
            if part.isdigit():
                ids_filter.append(int(part))

    flag_filter = request.args.get('flag', '').strip().upper()

    offset = (page - 1) * per_page

    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        if ids_filter and flag_filter:
            ph = ",".join(["%s"] * len(ids_filter))
            cursor.execute(
                f"""SELECT COUNT(DISTINCT r.person_id) AS total
                    FROM detection_records r
                    JOIN people p ON r.person_id = p.person_id
                    WHERE r.person_id IS NOT NULL AND r.person_id IN ({ph}) AND p.person_type_id = %s""",
                ids_filter + [flag_filter],
            )
        elif ids_filter:
            ph = ",".join(["%s"] * len(ids_filter))
            cursor.execute(
                f"SELECT COUNT(DISTINCT person_id) AS total FROM detection_records WHERE person_id IS NOT NULL AND person_id IN ({ph})",
                ids_filter,
            )
        elif flag_filter:
            cursor.execute(
                """SELECT COUNT(DISTINCT r.person_id) AS total
                   FROM detection_records r
                   JOIN people p ON r.person_id = p.person_id
                   WHERE r.person_id IS NOT NULL AND p.person_type_id = %s""",
                (flag_filter,),
            )
        else:
            cursor.execute("SELECT COUNT(DISTINCT person_id) AS total FROM detection_records WHERE person_id IS NOT NULL")
        total = cursor.fetchone()["total"]
        total_pages = max(1, (total + per_page - 1) // per_page)

        if ids_filter and flag_filter:
            ph = ",".join(["%s"] * len(ids_filter))
            cursor.execute(f"""
                SELECT r.person_id AS id_unico
                FROM detection_records r
                JOIN people p ON r.person_id = p.person_id
                WHERE r.person_id IS NOT NULL AND r.person_id IN ({ph}) AND p.person_type_id = %s
                GROUP BY r.person_id
                ORDER BY MAX(r.created_at) DESC
                LIMIT %s OFFSET %s
            """, ids_filter + [flag_filter, per_page, offset])
        elif ids_filter:
            ph = ",".join(["%s"] * len(ids_filter))
            cursor.execute(f"""
                SELECT person_id AS id_unico
                FROM detection_records
                WHERE person_id IS NOT NULL AND person_id IN ({ph})
                GROUP BY person_id
                ORDER BY MAX(created_at) DESC
                LIMIT %s OFFSET %s
            """, ids_filter + [per_page, offset])
        elif flag_filter:
            cursor.execute("""
                SELECT r.person_id AS id_unico
                FROM detection_records r
                JOIN people p ON r.person_id = p.person_id
                WHERE r.person_id IS NOT NULL AND p.person_type_id = %s
                GROUP BY r.person_id
                ORDER BY MAX(r.created_at) DESC
                LIMIT %s OFFSET %s
            """, (flag_filter, per_page, offset))
        else:
            cursor.execute("""
                SELECT person_id AS id_unico
                FROM detection_records
                WHERE person_id IS NOT NULL
                GROUP BY person_id
                ORDER BY MAX(created_at) DESC
                LIMIT %s OFFSET %s
            """, (per_page, offset))
        id_unicos = [row["id_unico"] for row in cursor.fetchall()]

        groups = []
        if id_unicos:
            placeholders = ",".join(["%s"] * len(id_unicos))
            cursor.execute(f"""
                SELECT
                    r.detection_record_id       AS id,
                    r.track_id,
                    r.person_id                 AS id_unico,
                    r.image_path,
                    r.created_at,
                    r.camera_id,
                    r.detection_score           AS face_det_score,
                    p.full_name                 AS nome,
                    p.nickname                  AS apelido,
                    p.person_type_id            AS flag,
                    p.gender_id                 AS genero,
                    p.reference_track_id        AS track_id_base,
                    p.document                  AS doc,
                    p.age                       AS idade,
                    p.notes                     AS notas
                FROM detection_records r
                LEFT JOIN people p ON r.person_id = p.person_id
                WHERE r.person_id IN ({placeholders})
                ORDER BY r.created_at DESC
            """, id_unicos)
            rows = cursor.fetchall()

            groups_dict = {}
            for row in rows:
                key = row["id_unico"]
                if key not in groups_dict:
                    groups_dict[key] = {
                        "id_unico":      key,
                        "nome":          row["nome"],
                        "apelido":       row["apelido"],
                        "flag":          row["flag"],
                        "genero":        row["genero"],
                        "doc":           row["doc"],
                        "idade":         row["idade"],
                        "notas":         row["notas"],
                        "track_id_base": row["track_id_base"],
                        "foto":          None,
                        "ocorrencias":   [],
                    }
                if groups_dict[key]["foto"] is None and row["image_path"] and row["track_id"] == groups_dict[key]["track_id_base"]:
                    groups_dict[key]["foto"] = HEIMDALL_IMAGE_BASE + row["image_path"]
                groups_dict[key]["ocorrencias"].append({
                    "id":             row["id"],
                    "track_id":       row["track_id"],
                    "image_url":      HEIMDALL_IMAGE_BASE + row["image_path"] if row["image_path"] else None,
                    "created_at":     fmt_timestamp(row["created_at"]),
                    "camera_id":      row["camera_id"],
                    "face_det_score": fmt_score(row["face_det_score"]) if row.get("face_det_score") is not None else None,
                })
            groups = [groups_dict[k] for k in id_unicos if k in groups_dict]

        cursor.close()
        conn.rollback()
    finally:
        release_faciais_conn(conn)

    cameras_map = {c["id_camera"]: c.get("camera") or str(c["id_camera"]) for c in CAMERAS_COMPLETO if c.get("id_camera") is not None}
    return render_template("tracks_lista.html", groups=groups, page=page, total_pages=total_pages,
                           cameras_map=cameras_map, ids_raw=ids_raw, flag_filter=flag_filter)


@tracks_bp.route("/tracks/permanencia")
def tracks_permanencia():
    per_page = 5
    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    offset = (page - 1) * per_page

    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cursor.execute("SELECT COUNT(DISTINCT person_id) AS total FROM detection_records WHERE person_id IS NOT NULL")
        total = cursor.fetchone()["total"]
        total_pages = max(1, (total + per_page - 1) // per_page)

        cursor.execute("""
            SELECT person_id AS id_unico
            FROM detection_records
            WHERE person_id IS NOT NULL
            GROUP BY person_id
            ORDER BY MAX(created_at) DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))
        id_unicos = [row["id_unico"] for row in cursor.fetchall()]

        groups = []
        if id_unicos:
            placeholders = ",".join(["%s"] * len(id_unicos))
            cursor.execute(f"""
                SELECT
                    r.detection_record_id       AS id,
                    r.track_id,
                    r.person_id                 AS id_unico,
                    r.image_path,
                    r.created_at,
                    p.full_name                 AS nome,
                    p.nickname                  AS apelido,
                    p.person_type_id            AS flag,
                    p.gender_id                 AS genero,
                    p.document                  AS doc,
                    p.age                       AS idade,
                    p.notes                     AS notas,
                    p.reference_track_id        AS track_id_base
                FROM detection_records r
                LEFT JOIN people p ON r.person_id = p.person_id
                WHERE r.person_id IN ({placeholders})
                ORDER BY r.created_at ASC
            """, id_unicos)
            rows = cursor.fetchall()

            groups_dict = {}
            for row in rows:
                key = row["id_unico"]
                dia = row["created_at"].date() if row["created_at"] else None

                if key not in groups_dict:
                    groups_dict[key] = {
                        "id_unico":      key,
                        "nome":          row["nome"],
                        "apelido":       row["apelido"],
                        "flag":          row["flag"],
                        "genero":        row["genero"],
                        "doc":           row["doc"],
                        "idade":         row["idade"],
                        "notas":         row["notas"],
                        "track_id_base": row["track_id_base"],
                        "foto":          None,
                        "dias":          {},
                    }

                g = groups_dict[key]
                if g["foto"] is None and row["image_path"] and row["track_id"] == g["track_id_base"]:
                    g["foto"] = HEIMDALL_IMAGE_BASE + row["image_path"]

                if dia not in g["dias"]:
                    g["dias"][dia] = []
                g["dias"][dia].append(row)

            for key, g in groups_dict.items():
                dias_list = []
                for dia in sorted(g["dias"].keys(), reverse=True):
                    regs = g["dias"][dia]
                    primeira = regs[0]
                    ultima = regs[-1]

                    if len(regs) == 1 or (ultima["created_at"] - primeira["created_at"]) < timedelta(minutes=2):
                        duracao = timedelta(minutes=30)
                        estimado = True
                    else:
                        duracao = ultima["created_at"] - primeira["created_at"]
                        estimado = False

                    total_secs = int(duracao.total_seconds())
                    horas = total_secs // 3600
                    mins = (total_secs % 3600) // 60
                    if horas > 0:
                        perm_str = f"{horas}h {mins:02d}min"
                    else:
                        perm_str = f"{mins}min"

                    dias_list.append({
                        "dia":       str(dia),
                        "hora_ini":  fmt_timestamp(primeira["created_at"]),
                        "hora_fim":  fmt_timestamp(ultima["created_at"]) if not estimado else None,
                        "permanencia": perm_str,
                        "estimado":  estimado,
                        "image_url": HEIMDALL_IMAGE_BASE + primeira["image_path"] if primeira["image_path"] else None,
                    })
                g["dias"] = dias_list

            groups = [groups_dict[k] for k in id_unicos if k in groups_dict]

        cursor.close()
        conn.rollback()
    finally:
        release_faciais_conn(conn)

    return render_template("tracks_permanencia.html", groups=groups, page=page, total_pages=total_pages)


# Mapeamento de campos do frontend (nomes legados) para colunas do schema faciais
_CAMPO_MAP = {
    "nome":    "full_name",
    "apelido": "nickname",
    "doc":     "document",
    "idade":   "age",
    "genero":  "gender_id",
    "flag":    "person_type_id",
    "notas":   "notes",
}


@tracks_bp.route("/tracks/api/pessoa/<int:id_unico>", methods=["POST"])
def atualizar_pessoa(id_unico):
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON inválido"}), 400
    updates = {_CAMPO_MAP[k]: (v if v != "" else None) for k, v in data.items() if k in _CAMPO_MAP}
    if not updates:
        return jsonify({"error": "Nenhum campo válido"}), 400
    set_clause = ", ".join(f"{col} = %s" for col in updates)
    values = list(updates.values()) + [id_unico]
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor()
        cursor.execute(f"UPDATE people SET {set_clause} WHERE person_id = %s", values)
        conn.commit()
        cursor.close()
    except Exception:
        conn.rollback()
        raise
    finally:
        release_faciais_conn(conn)
    return jsonify({"success": True})


@tracks_bp.route("/tracks/api/pessoa/<int:id_unico>/base", methods=["POST"])
def atualizar_base_pessoa(id_unico):
    data = request.get_json(silent=True)
    if not data or "track_id_base" not in data:
        return jsonify({"error": "track_id_base obrigatório"}), 400
    track_id_base = data["track_id_base"]
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE people SET reference_track_id = %s WHERE person_id = %s", (track_id_base, id_unico))
        conn.commit()
        cursor.close()
    except Exception:
        conn.rollback()
        raise
    finally:
        release_faciais_conn(conn)
    return jsonify({"success": True})


@tracks_bp.route("/tracks/api/pessoa/<int:id_unico>", methods=["DELETE"])
def excluir_pessoa(id_unico):
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM detection_records WHERE person_id = %s", (id_unico,))
        cursor.execute("DELETE FROM people WHERE person_id = %s", (id_unico,))
        conn.commit()
        cursor.close()
    except Exception:
        conn.rollback()
        raise
    finally:
        release_faciais_conn(conn)
    return jsonify({"success": True})


@tracks_bp.route("/tracks/api/pessoa/<int:id_unico>", methods=["GET"])
def buscar_pessoa(id_unico):
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            SELECT p.person_id AS id_unico, p.full_name AS nome, p.nickname AS apelido,
                   p.person_type_id AS flag,
                   (SELECT r.image_path FROM detection_records r
                    WHERE r.track_id = p.reference_track_id AND r.image_path IS NOT NULL
                    ORDER BY r.detection_score DESC LIMIT 1) AS image_path
            FROM people p
            WHERE p.person_id = %s
            """,
            (id_unico,),
        )
        pessoa = cursor.fetchone()
        cursor.close()
        conn.rollback()
    finally:
        release_faciais_conn(conn)
    if pessoa is None:
        return jsonify({"error": f"Pessoa {id_unico} não encontrada"}), 404
    pessoa = dict(pessoa)
    pessoa["foto"] = HEIMDALL_IMAGE_BASE + pessoa["image_path"] if pessoa.get("image_path") else None
    del pessoa["image_path"]
    return jsonify(pessoa)


@tracks_bp.route("/tracks/api/registro/<int:reg_id>", methods=["DELETE"])
def excluir_registro(reg_id):
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM detection_records WHERE detection_record_id = %s", (reg_id,))
        conn.commit()
        cursor.close()
    except Exception:
        conn.rollback()
        raise
    finally:
        release_faciais_conn(conn)
    return jsonify({"success": True})


@tracks_bp.route("/tracks/api/registro/<int:reg_id>/mover", methods=["POST"])
def mover_registro(reg_id):
    data = request.get_json(silent=True)
    if not data or "id_unico" not in data:
        return jsonify({"error": "id_unico obrigatório"}), 400
    id_unico = data["id_unico"]
    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute("SELECT person_id FROM people WHERE person_id = %s", (id_unico,))
        if cursor.fetchone() is None:
            cursor.close()
            release_faciais_conn(conn)
            return jsonify({"error": f"Pessoa {id_unico} não encontrada"}), 404

        cursor.execute("SELECT person_id AS id_unico FROM detection_records WHERE detection_record_id = %s", (reg_id,))
        row = cursor.fetchone()
        id_unico_original = row["id_unico"] if row else None

        cursor.execute("UPDATE detection_records SET person_id = %s WHERE detection_record_id = %s", (id_unico, reg_id))

        pessoa_excluida = False
        if id_unico_original and id_unico_original != id_unico:
            cursor.execute("SELECT COUNT(*) AS total FROM detection_records WHERE person_id = %s", (id_unico_original,))
            if cursor.fetchone()["total"] == 0:
                cursor.execute("DELETE FROM people WHERE person_id = %s", (id_unico_original,))
                pessoa_excluida = True

        conn.commit()
        cursor.close()
    except Exception:
        conn.rollback()
        raise
    finally:
        release_faciais_conn(conn)
    return jsonify({"success": True, "pessoa_excluida": pessoa_excluida})


@tracks_bp.route("/tracks/tabuleiro")
def tracks_tabuleiro():
    per_page = 30
    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    offset = (page - 1) * per_page

    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cursor.execute("SELECT COUNT(*) AS total FROM people")
        total = cursor.fetchone()["total"]
        total_pages = max(1, (total + per_page - 1) // per_page)

        cursor.execute("""
            SELECT
                p.person_id                 AS id_unico,
                p.full_name                 AS nome,
                p.nickname                  AS apelido,
                p.person_type_id            AS flag,
                p.gender_id                 AS genero,
                p.document                  AS doc,
                p.age                       AS idade,
                p.notes                     AS notas,
                (SELECT r.image_path FROM detection_records r
                 WHERE r.track_id = p.reference_track_id AND r.image_path IS NOT NULL
                 ORDER BY r.detection_score DESC LIMIT 1) AS image_path
            FROM people p
            ORDER BY p.person_id DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))
        pessoas = cursor.fetchall()

        cursor.close()
        conn.rollback()
    finally:
        release_faciais_conn(conn)

    pessoas = [dict(p) for p in pessoas]
    for p in pessoas:
        p["foto"] = HEIMDALL_IMAGE_BASE + p["image_path"] if p["image_path"] else None

    return render_template("tracks_tabuleiro.html", pessoas=pessoas, page=page, total_pages=total_pages)


@tracks_bp.route("/tracks/quadro")
def tracks_quadro():
    from datetime import date, timedelta
    today = date.today()
    week_ago = today - timedelta(days=7)

    d_ini = request.args.get('d_ini', week_ago.isoformat())
    d_fim = request.args.get('d_fim', today.isoformat())
    h_ini = request.args.get('h_ini', week_ago.isoformat())
    h_fim = request.args.get('h_fim', today.isoformat())
    p_ini = request.args.get('p_ini', week_ago.isoformat())
    p_fim = request.args.get('p_fim', today.isoformat())
    e_ini = request.args.get('e_ini', week_ago.isoformat())
    e_fim = request.args.get('e_fim', today.isoformat())
    perm_ini = request.args.get('perm_ini', week_ago.isoformat())
    perm_fim = request.args.get('perm_fim', today.isoformat())
    gd_ini = request.args.get('gd_ini', week_ago.isoformat())
    gd_fim = request.args.get('gd_fim', today.isoformat())
    fat_ini = request.args.get('fat_ini', week_ago.isoformat())
    fat_fim = request.args.get('fat_fim', today.isoformat())
    top_ini = request.args.get('top_ini', week_ago.isoformat())
    top_fim = request.args.get('top_fim', today.isoformat())

    # ── Faturamento diário (PostgreSQL microvix) ──────────────────────────────
    faturamento_por_dia = []
    pg_conn = None
    try:
        pg_conn = get_pg_conn()
        pg_cur = pg_conn.cursor()
        pg_cur.execute("""
            SELECT
                data_lancamento::date AS data_doc,
                COUNT(DISTINCT documento) AS qtd_notas,
                SUM(valor_liquido) AS soma
            FROM microvix_movimento
            WHERE cod_natureza_operacao = '10030'
              AND cancelado = 'N'
              AND excluido = 'N'
              AND tipo_transacao = 'V'
              AND data_lancamento::date BETWEEN %s AND %s
            GROUP BY data_lancamento::date
            ORDER BY data_lancamento::date ASC
        """, (fat_ini, fat_fim))
        for row in pg_cur.fetchall():
            faturamento_por_dia.append({
                "dia": str(row[0]),
                "qtd_notas": int(row[1]),
                "soma": float(row[2]) if row[2] is not None else 0.0,
                "qtd_clientes": 0,
            })
        pg_cur.close()
    except Exception as e:
        print(f"[quadro] Erro PostgreSQL faturamento: {e}")
    finally:
        if pg_conn:
            release_pg_conn(pg_conn)

    # ── Top 10 produtos por faturamento (PostgreSQL microvix) ─────────────────
    top10_produtos = []
    pg_conn2 = None
    try:
        pg_conn2 = get_pg_conn()
        pg_cur2 = pg_conn2.cursor()
        pg_cur2.execute("""
            SELECT
                m.cod_produto,
                p.nome AS nome_produto,
                COUNT(DISTINCT m.documento) AS qtd_notas,
                ROUND(SUM(m.valor_liquido)::numeric, 2) AS soma_valor
            FROM microvix_movimento m
            LEFT JOIN microvix_produtos p
                   ON m.cod_produto = p.cod_produto
            WHERE m.cod_natureza_operacao = '10030'
              AND m.cancelado = 'N'
              AND m.excluido = 'N'
              AND m.tipo_transacao = 'V'
              AND m.data_lancamento::date BETWEEN %s AND %s
            GROUP BY m.cod_produto, p.nome
            ORDER BY soma_valor DESC
            LIMIT 10
        """, (top_ini, top_fim))
        for row in pg_cur2.fetchall():
            top10_produtos.append({
                "nome": row[1] or f"Cód. {row[0]}",
                "qtd_notas": int(row[2]),
                "soma_valor": float(row[3]) if row[3] is not None else 0.0,
            })
        pg_cur2.close()
    except Exception as e:
        print(f"[quadro] Erro PostgreSQL top10 produtos: {e}")
    finally:
        if pg_conn2:
            release_pg_conn(pg_conn2)

    # ── Clientes por dia no intervalo do faturamento (PostgreSQL faciais) ─────
    if faturamento_por_dia:
        try:
            _conn_fat = get_faciais_conn()
            try:
                _cur_fat = _conn_fat.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                _cur_fat.execute("""
                    SELECT DATE(r.created_at) AS dia, COUNT(DISTINCT r.person_id) AS total
                    FROM detection_records r
                    JOIN people p ON r.person_id = p.person_id
                    WHERE DATE(r.created_at) BETWEEN %s AND %s
                      AND p.person_type_id IN ('C', 'A')
                    GROUP BY DATE(r.created_at)
                """, (fat_ini, fat_fim))
                clientes_map = {str(row["dia"]): row["total"] for row in _cur_fat.fetchall()}
                _cur_fat.close()
                _conn_fat.rollback()
                for entry in faturamento_por_dia:
                    entry["qtd_clientes"] = clientes_map.get(entry["dia"], 0)
            finally:
                release_faciais_conn(_conn_fat)
        except Exception as e:
            print(f"[quadro] Erro faciais clientes/faturamento: {e}")

    conn = get_faciais_conn()
    try:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cursor.execute("""
            SELECT
                DATE(r.created_at) AS dia,
                COUNT(DISTINCT CASE WHEN mg.primeira = DATE(r.created_at) THEN r.person_id END) AS novos,
                COUNT(DISTINCT CASE WHEN mg.primeira < DATE(r.created_at) THEN r.person_id END) AS retornantes
            FROM detection_records r
            JOIN people p ON r.person_id = p.person_id
            JOIN (
                SELECT person_id, MIN(DATE(created_at)) AS primeira
                FROM detection_records
                GROUP BY person_id
            ) mg ON r.person_id = mg.person_id
            WHERE DATE(r.created_at) BETWEEN %s AND %s
              AND p.person_type_id IN ('C', 'A')
            GROUP BY DATE(r.created_at)
            ORDER BY dia ASC
        """, (d_ini, d_fim))
        rows = cursor.fetchall()
        ocorrencias_por_dia = [
            {"dia": str(r["dia"]), "novos": r["novos"], "retornantes": r["retornantes"]}
            for r in rows
        ]

        cursor.execute("""
            SELECT EXTRACT(HOUR FROM primeira.created_at)::int AS hora, COUNT(*) AS total
            FROM (
                SELECT r.person_id, DATE(r.created_at) AS dia, MIN(r.created_at) AS created_at
                FROM detection_records r
                JOIN people p ON r.person_id = p.person_id
                WHERE DATE(r.created_at) BETWEEN %s AND %s
                  AND p.person_type_id IN ('C', 'A')
                GROUP BY r.person_id, DATE(r.created_at)
            ) AS primeira
            GROUP BY EXTRACT(HOUR FROM primeira.created_at)::int
            ORDER BY hora ASC
        """, (h_ini, h_fim))
        rows_hora = cursor.fetchall()

        cursor.execute("""
            SELECT p.nickname AS apelido, COUNT(DISTINCT DATE(r.created_at)) AS total
            FROM detection_records r
            JOIN people p ON r.person_id = p.person_id
            WHERE DATE(r.created_at) BETWEEN %s AND %s
              AND p.person_type_id IN ('C', 'A')
            GROUP BY r.person_id, p.nickname
            HAVING COUNT(DISTINCT DATE(r.created_at)) >= 3
            ORDER BY total DESC
        """, (p_ini, p_fim))
        rows_pessoa = cursor.fetchall()

        cursor.execute("""
            SELECT
              CASE
                WHEN p.age BETWEEN 0  AND 25 THEN '00-25'
                WHEN p.age BETWEEN 26 AND 35 THEN '26-35'
                WHEN p.age BETWEEN 36 AND 45 THEN '36-45'
                WHEN p.age BETWEEN 46 AND 55 THEN '46-55'
                WHEN p.age BETWEEN 56 AND 65 THEN '56-65'
                WHEN p.age > 65              THEN '65+'
              END AS faixa,
              COUNT(*) AS total
            FROM (
                SELECT DISTINCT r.person_id, DATE(r.created_at) AS dia
                FROM detection_records r
                JOIN people p ON r.person_id = p.person_id
                WHERE DATE(r.created_at) BETWEEN %s AND %s
                  AND p.person_type_id = 'C'
                  AND p.age IS NOT NULL
            ) AS uniq
            JOIN people p ON uniq.person_id = p.person_id
            GROUP BY faixa
        """, (e_ini, e_fim))
        rows_etaria = cursor.fetchall()

        cursor.execute("""
            SELECT r.person_id AS id_unico, p.gender_id AS genero,
                   MIN(r.created_at) AS primeira, MAX(r.created_at) AS ultima
            FROM detection_records r
            JOIN people p ON r.person_id = p.person_id
            WHERE DATE(r.created_at) BETWEEN %s AND %s
              AND p.person_type_id IN ('C', 'A')
            GROUP BY r.person_id, DATE(r.created_at), p.gender_id
        """, (perm_ini, perm_fim))
        rows_perm = cursor.fetchall()

        cursor.execute("""
            SELECT DATE(r.created_at) AS dia, p.gender_id AS genero, COUNT(DISTINCT r.person_id) AS total
            FROM detection_records r
            JOIN people p ON r.person_id = p.person_id
            WHERE DATE(r.created_at) BETWEEN %s AND %s
              AND p.person_type_id IN ('C', 'A')
              AND p.gender_id IN ('M', 'F')
            GROUP BY DATE(r.created_at), p.gender_id
            ORDER BY dia ASC
        """, (gd_ini, gd_fim))
        rows_genero_dia = cursor.fetchall()

        cursor.close()
        conn.rollback()
    finally:
        release_faciais_conn(conn)

    def _perm_minutos(primeira, ultima):
        diff = ultima - primeira
        if diff < timedelta(minutes=2):
            return 30.0
        return diff.total_seconds() / 60

    def _fmt_media(vals):
        if not vals:
            return "—"
        avg = sum(vals) / len(vals)
        h = int(avg) // 60
        m = int(avg) % 60
        return f"{h}h {m:02d}min" if h > 0 else f"{m}min"

    perms_total, perms_M, perms_F = [], [], []
    for row in rows_perm:
        mins = _perm_minutos(row["primeira"], row["ultima"])
        perms_total.append(mins)
        if row["genero"] == "M":
            perms_M.append(mins)
        elif row["genero"] == "F":
            perms_F.append(mins)

    def _avg(vals):
        return round(sum(vals) / len(vals), 1) if vals else 0

    permanencia_media = [
        {"label": "Total",     "minutos": _avg(perms_total), "fmt": _fmt_media(perms_total)},
        {"label": "Masculino", "minutos": _avg(perms_M),     "fmt": _fmt_media(perms_M)},
        {"label": "Feminino",  "minutos": _avg(perms_F),     "fmt": _fmt_media(perms_F)},
    ]

    gd_map = {}
    for row in rows_genero_dia:
        dia = str(row["dia"])
        if dia not in gd_map:
            gd_map[dia] = {"M": 0, "F": 0}
        gd_map[dia][row["genero"]] = row["total"]
    genero_por_dia = [
        {"dia": d, "M": gd_map[d]["M"], "F": gd_map[d]["F"]}
        for d in sorted(gd_map)
    ]

    horas_map = {r["hora"]: r["total"] for r in rows_hora}
    ocorrencias_por_hora = [
        {"faixa": "00h00-07h59", "total": sum(horas_map.get(h, 0) for h in range(0, 8))},
        *[{"faixa": f"{h:02d}h00-{h:02d}h59", "total": horas_map.get(h, 0)} for h in range(8, 19)],
        {"faixa": "19h00-21h59", "total": sum(horas_map.get(h, 0) for h in range(19, 22))},
        {"faixa": "22h00-23h59", "total": sum(horas_map.get(h, 0) for h in range(22, 24))},
    ]
    ocorrencias_por_pessoa = [
        {"apelido": r["apelido"], "total": r["total"]}
        for r in rows_pessoa
    ]
    faixas_ordem = ['00-25', '26-35', '36-45', '46-55', '56-65', '65+']
    etaria_map = {r["faixa"]: r["total"] for r in rows_etaria}
    ocorrencias_por_faixa_etaria = [
        {"faixa": f, "total": etaria_map.get(f, 0)}
        for f in faixas_ordem
    ]

    return render_template("tracks_quadro.html",
                           ocorrencias_por_dia=ocorrencias_por_dia,
                           ocorrencias_por_hora=ocorrencias_por_hora,
                           ocorrencias_por_pessoa=ocorrencias_por_pessoa,
                           ocorrencias_por_faixa_etaria=ocorrencias_por_faixa_etaria,
                           permanencia_media=permanencia_media,
                           faturamento_por_dia=faturamento_por_dia,
                           top10_produtos=top10_produtos,
                           d_ini=d_ini, d_fim=d_fim,
                           h_ini=h_ini, h_fim=h_fim,
                           p_ini=p_ini, p_fim=p_fim,
                           e_ini=e_ini, e_fim=e_fim,
                           perm_ini=perm_ini, perm_fim=perm_fim,
                           genero_por_dia=genero_por_dia,
                           gd_ini=gd_ini, gd_fim=gd_fim,
                           fat_ini=fat_ini, fat_fim=fat_fim,
                           top_ini=top_ini, top_fim=top_fim)


@tracks_bp.route("/tracks/dados")
def tracks_dados():
    _, cameras = _carregar_cameras()
    return render_template("tracks_dados.html", cameras=cameras)


@tracks_bp.route("/tracks/snapshot/<int:camera_id>")
def tracks_snapshot(camera_id):
    cam = next((c for c in CAMERAS_COMPLETO if c.get("id_camera") == camera_id), None)
    if cam is None or not cam.get("rstp"):
        return Response("Câmera não encontrada ou sem URL RTSP.", status=404, mimetype="text/plain")

    rtsp_url = cam["rstp"]
    cap = cv2.VideoCapture(rtsp_url)
    try:
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        ret, frame = cap.read()
        if not ret or frame is None:
            return Response("Não foi possível capturar frame.", status=502, mimetype="text/plain")
        ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
        if not ok:
            return Response("Erro ao codificar imagem.", status=500, mimetype="text/plain")
        return Response(io.BytesIO(buf.tobytes()).read(), mimetype="image/jpeg")
    finally:
        cap.release()


@tracks_bp.route("/tracks/api")
def tracks_api():
    rows = get_last_track_ids(5)
    results = []
    for row in rows:
        track_id = row["track_id"]
        data, error = query_heimdall(track_id)
        results.append({
            "track_id": track_id,
            "data": data,
            "error": error,
        })
    return jsonify(results)


_GENERO_LABEL = {"M": "Masculino", "F": "Feminino", "A": "Anônimo"}
_FLAG_LABEL   = {"C": "Cliente",   "A": "Anônimo",  "F": "Franqueado", "E": "Empregado", "K": "Criança", "P": "Prestador"}


@tracks_bp.route("/tracks/export")
def tracks_export():
    today     = date.today()
    week_ago  = today - timedelta(days=7)
    data_ini  = request.args.get("data_ini", week_ago.isoformat())
    data_fim  = request.args.get("data_fim", today.isoformat())
    return render_template("tracks_export.html", data_ini=data_ini, data_fim=data_fim)


@tracks_bp.route("/tracks/export/download")
def tracks_export_download():
    today    = date.today()
    week_ago = today - timedelta(days=7)
    data_ini = request.args.get("data_ini", week_ago.isoformat())
    data_fim = request.args.get("data_fim", today.isoformat())

    try:
        conn = get_faciais_conn()
        try:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute("""
                SELECT
                    r.detection_record_id       AS id,
                    r.camera_id,
                    r.created_at,
                    r.person_id                 AS id_unico,
                    p.full_name                 AS nome,
                    p.nickname                  AS apelido,
                    p.age                       AS idade,
                    p.gender_id                 AS genero,
                    p.person_type_id            AS flag,
                    tc.camera_type_name         AS tipo_camera,
                    c.camera_name               AS camera,
                    e.company_name              AS empresa
                FROM detection_records r
                LEFT JOIN people        p  ON p.person_id         = r.person_id
                LEFT JOIN cameras       c  ON c.camera_id          = r.camera_id
                LEFT JOIN camera_types  tc ON tc.camera_type_id    = c.camera_type_id
                LEFT JOIN stores        l  ON l.store_id           = c.store_id
                LEFT JOIN companies     e  ON e.company_id         = l.company_id
                WHERE DATE(r.created_at) BETWEEN %s AND %s
                ORDER BY r.created_at
            """, (data_ini, data_fim))
            rows = cursor.fetchall()
            cursor.close()
            conn.rollback()
        finally:
            release_faciais_conn(conn)
    except Exception as e:
        print(f"[export] Erro no banco: {e}")
        return Response(f"Erro ao consultar banco de dados: {e}", status=500, mimetype="text/plain")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"

        header_fill = PatternFill(patternType="solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Ocor", "Id Cam", "DataHora", "Id Pessoa",
                   "Nome", "Apelido", "Idade", "Gênero", "Flag",
                   "Tipo Câmera", "Nome Câmera", "Empresa"]
        col_widths = [8, 8, 20, 10, 25, 15, 8, 12, 12, 15, 20, 25]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1,  value=row["id"])
            ws.cell(row=row_idx, column=2,  value=row["camera_id"])
            ws.cell(row=row_idx, column=3,  value=row["created_at"].strftime("%d/%m/%Y %H:%M:%S") if row["created_at"] else "")
            ws.cell(row=row_idx, column=4,  value=row["id_unico"])
            ws.cell(row=row_idx, column=5,  value=row["nome"] or "")
            ws.cell(row=row_idx, column=6,  value=row["apelido"] or "")
            ws.cell(row=row_idx, column=7,  value=row["idade"])
            ws.cell(row=row_idx, column=8,  value=_GENERO_LABEL.get(row["genero"], row["genero"] or ""))
            ws.cell(row=row_idx, column=9,  value=_FLAG_LABEL.get(row["flag"], row["flag"] or ""))
            ws.cell(row=row_idx, column=10, value=row["tipo_camera"] or "")
            ws.cell(row=row_idx, column=11, value=row["camera"] or "")
            ws.cell(row=row_idx, column=12, value=row["empresa"] or "")

        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
    except Exception as e:
        print(f"[export] Erro ao gerar planilha: {e}")
        return Response(f"Erro ao gerar planilha: {e}", status=500, mimetype="text/plain")

    filename = f"registros_{data_ini}_{data_fim}.xlsx"
    resp = make_response(xlsx_bytes)
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Content-Length"] = len(xlsx_bytes)
    return resp


@tracks_bp.route("/tracks/caixa")
def tracks_caixa():
    from datetime import timedelta

    # 1. Buscar as 25 últimas notas fiscais (PostgreSQL microvix)
    notas = []
    pg_conn = None
    try:
        pg_conn = get_pg_conn()
        pg_cur = pg_conn.cursor()
        pg_cur.execute("""
            SELECT
                documento,
                COUNT(*)                                       AS itens,
                ROUND(SUM(valor_liquido)::numeric, 2)            AS valor,
                data_lancamento::date                          AS data,
                hora_lancamento                                AS hora,
                (data_lancamento::date + hora_lancamento::time) AS nf_dt
            FROM microvix_movimento
            WHERE cod_natureza_operacao = '10030'
              AND cancelado = 'N'
              AND excluido = 'N'
              AND tipo_transacao = 'V'
            GROUP BY documento, data_lancamento::date, hora_lancamento
            ORDER BY data_lancamento::date DESC, hora_lancamento DESC
            LIMIT 25
        """)
        for row in pg_cur.fetchall():
            notas.append({
                "documento": row[0],
                "itens":     int(row[1]),
                "valor":     float(row[2]) if row[2] is not None else 0.0,
                "data":      row[3],
                "hora":      (row[4] or "").strip(),
                "nf_dt":     row[5],
                "clientes":  [],
            })
        pg_cur.close()
    except Exception as e:
        print(f"[caixa] Erro PostgreSQL: {e}")
    finally:
        if pg_conn:
            release_pg_conn(pg_conn)

    # 2. Buscar registros faciais (câmera Caixa, person_type_id='C') no intervalo total ±2 min
    if notas:
        dts = [n["nf_dt"] for n in notas if n["nf_dt"] is not None]
        if dts:
            dt_min = min(dts) - timedelta(minutes=1)
            dt_max = max(dts) + timedelta(minutes=1)
            try:
                conn = get_faciais_conn()
                try:
                    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                    cursor.execute("""
                        SELECT
                            r.detection_record_id   AS id,
                            r.person_id             AS id_unico,
                            r.image_path,
                            r.created_at,
                            p.full_name             AS nome,
                            p.nickname              AS apelido,
                            p.gender_id             AS genero,
                            p.age                   AS idade
                        FROM detection_records r
                        JOIN cameras c        ON c.camera_id       = r.camera_id
                        JOIN camera_types tc  ON tc.camera_type_id = c.camera_type_id
                        JOIN people p         ON p.person_id        = r.person_id
                        WHERE LOWER(tc.camera_type_name) = 'caixa'
                          AND p.person_type_id = 'C'
                          AND r.created_at BETWEEN %s AND %s
                        ORDER BY r.created_at DESC
                    """, (dt_min, dt_max))
                    registros = cursor.fetchall()
                    cursor.close()
                    conn.rollback()
                finally:
                    release_faciais_conn(conn)

                registros = [dict(r) for r in registros]
                for reg in registros:
                    reg["foto"] = HEIMDALL_IMAGE_BASE + reg["image_path"] if reg["image_path"] else None

                for nota in notas:
                    if nota["nf_dt"] is None:
                        continue
                    janela_ini = nota["nf_dt"] - timedelta(minutes=1)
                    janela_fim = nota["nf_dt"] + timedelta(minutes=1)
                    seen = set()
                    for reg in registros:
                        if janela_ini <= reg["created_at"] <= janela_fim:
                            if reg["id_unico"] not in seen:
                                seen.add(reg["id_unico"])
                                nota["clientes"].append(reg)
            except Exception as e:
                print(f"[caixa] Erro faciais: {e}")

    return render_template("tracks_caixa.html", notas=notas)


@tracks_bp.route("/tracks/caixa/nf/<documento>")
def tracks_caixa_nf_itens(documento):
    pg_conn = None
    try:
        pg_conn = get_pg_conn()
        pg_cur = pg_conn.cursor()
        pg_cur.execute("""
            SELECT
                m.cod_produto,
                p.nome AS produto_nome,
                m.quantidade,
                m.valor_liquido
            FROM microvix_movimento m
            LEFT JOIN microvix_produtos p ON m.cod_produto = p.cod_produto
            WHERE m.documento = %s
              AND m.cod_natureza_operacao = '10030'
              AND m.cancelado = 'N'
              AND m.excluido = 'N'
              AND m.tipo_transacao = 'V'
            ORDER BY m.cod_produto
        """, (documento,))
        rows = pg_cur.fetchall()
        pg_cur.close()
    except Exception as e:
        print(f"[caixa/nf] Erro PostgreSQL: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if pg_conn:
            release_pg_conn(pg_conn)

    itens = [
        {
            "cod_produto":    row[0],
            "produto_nome":   row[1] or "—",
            "quantidade":     float(row[2]) if row[2] is not None else 0,
            "valor_unitario": float(row[3]) if row[3] is not None else 0.0,
        }
        for row in rows
    ]
    return jsonify(itens)


@tracks_bp.route("/tracks/carga")
def tracks_carga():
    cargas = []
    pg_conn = None
    try:
        pg_conn = get_pg_conn()
        pg_cur = pg_conn.cursor()
        pg_cur.execute("""
            SELECT
                id_carga,
                microvix_grupo_lojas,
                microvix_lojas,
                microvix_clientes_fornecedores,
                microvix_movimento,
                microvix_produtos,
                microvix_produtos_detalhes,
                created_at
            FROM microvix_carga
            ORDER BY created_at DESC
            LIMIT 20
        """)
        for row in pg_cur.fetchall():
            cargas.append({
                "id_carga":                      row[0],
                "microvix_grupo_lojas":           row[1],
                "microvix_lojas":                 row[2],
                "microvix_clientes_fornecedores": row[3],
                "microvix_movimento":             row[4],
                "microvix_produtos":              row[5],
                "microvix_produtos_detalhes":     row[6],
                "created_at":                     row[7],
            })
        pg_cur.close()
    except Exception as e:
        print(f"[carga] Erro PostgreSQL: {e}")
    finally:
        if pg_conn:
            release_pg_conn(pg_conn)
    return render_template("tracks_carga.html", cargas=cargas)


@tracks_bp.route("/tracks/carga/sync", methods=["POST"])
def tracks_carga_sync():
    import threading as _threading
    from microvix_ingest import run_incremental, get_status
    st = get_status()
    if st["running"]:
        return jsonify({"ok": False, "error": "Sincronização já em andamento."})
    t = _threading.Thread(target=run_incremental, daemon=True)
    t.start()
    return jsonify({"ok": True})


@tracks_bp.route("/tracks/carga/status")
def tracks_carga_status():
    from microvix_ingest import get_status
    return jsonify(get_status())


@tracks_bp.route("/tracks/logs")
def tracks_logs():
    error = None
    logs = []
    try:
        resp = requests.get(
            f"{ZIONS_API_URL}/facial_recognition_log",
            headers={
                "Accept": "application/json",
                "Authorization": f"Bearer {ZIONS_TOKEN}",
                "paginate": "10",
                "orderDirection": "desc",
            },
            timeout=10,
        )
        if resp.ok:
            body = resp.json()
            logs = body.get("data", [])
        else:
            error = f"HTTP {resp.status_code}: {resp.text[:300]}"
    except Exception as e:
        error = str(e)
    return render_template("tracks_logs.html", logs=logs, error=error)
